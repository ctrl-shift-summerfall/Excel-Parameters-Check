# OS libraries import list:
import datetime
from pprint import pprint
import sys
import os

# Excel managers import list:
from openpyxl.styles import PatternFill 
import openpyxl

# Typing and classes import list:
from dataclasses import asdict, dataclass
from typing import Any
import re

# PyQt6 Interface import list:
from PyQt6.QtWidgets import *
from PyQt6.QtGui import QFont
from PyQt6 import QtCore


@dataclass(frozen=True, order=True)
class Cell:
    position: str
    value: Any


class Workbook:

    def __init__(self):

        self._default_workbook_file_name: str = 'Book1'
        self._default_workbook_file_extension: str = 'xlsx'
        self._default_workbook_file_location: str = ''
        self._default_worksheet_name_param: str = 'P_LIST'
        self._default_worksheet_name_results: str = 'P_RESULT'

        self._save_prefix_allowed: bool = False
        self._save_prefix: str = None
        self._save_postfix_allowed: bool = False
        self._save_postfix: str = None
        self._save_timestamp_allowed: bool = False
        self._save_timestamp_format: str = '%Y%m%d_%H%M%S'

        self._active_workbook: openpyxl.Workbook = None
        self._active_workbook_name: str = None
        self._active_workbook_path: str = None
        self._active_workbook_worksheet_list: list = None
        self._active_worksheet: openpyxl.Workbook[str] = None
        self._active_worksheet_name: str = None

        self._reserved_row: int = 2

        
    @classmethod
    def get_column_list(self):
        
        # Generating column indexed list:
        column_list = []
        column_index = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        column_index_depth_max = int(26 - 1)
        depth_current = -1
        while depth_current <= column_index_depth_max:
            for letter in column_index:
                column = letter
                if depth_current >= 0:
                    prefix = column_index[depth_current - 1]
                    column = f'{prefix}{letter}'
                column_list.append(column)
            depth_current += 1
        
        # Sorting columns:
        column_list.sort()
        column_list_sorted = []
        for column in column_list:
            if len(column) == 1:
                column_list_sorted.append(column)
        for column in column_list:
            if len(column) == 2:
                column_list_sorted.append(column)
        column_list = column_list_sorted
        
        # Returning column list:
        return column_list

    def open_workbook(self, document_path: str):
        if os.path.exists(document_path):

            # Handling workbook object:
            workbook_object = openpyxl.load_workbook(filename=document_path)
            self._active_workbook = workbook_object
            # self._active_workbook_name = TODO: Regex
            self._active_workbook_path = document_path
            self._active_workbook_worksheet_list = workbook_object.sheetnames

            # Handling worksheet object:
            worksheet_name = self._active_workbook_worksheet_list[0]
            self._active_worksheet = self._active_workbook[worksheet_name]
            self._active_worksheet_name = worksheet_name

    def close_workbook(self):

        # Resetting class attributes:
        self._active_workbook: openpyxl.Workbook = None
        self._active_workbook_name: str = None
        self._active_workbook_path: str = None
        self._active_workbook_worksheet_list: list = None
        self._active_worksheet: openpyxl.Workbook[str] = None
        self._active_worksheet_name: str = None

        pass

    def save_workbook(self,):
        pass

    def save_workbook_as(self):
        
        def generate_timestamp():

            # %Y%m%d_%H%M%S (Default) -> 20221027_123501
            datetime_now = datetime.datetime.now()
            datetime_timestamp_format: str = self._save_timestamp_format
            datetime_timestamp: str = datetime.datetime.strftime(datetime_now, datetime_timestamp_format)

            return datetime_timestamp

        def generate_save_filename():

            # PREFIX_NAME_POSTFIX_TIMEXTAMP.EXTENSION -> checked_book1_custom_20221027_123501.xlsx
            save_filename = '{set_prefix}{file_name}{set_postfix}{set_timestamp}.{file_extension}'.format(
                set_prefix = f'{self._save_prefix}_' if self._save_prefix_allowed else '',
                set_postfix = f'_{self._save_postfix}' if self._save_postfix_allowed else '',
                set_timestamp = f'_{generate_timestamp()}' if self._save_timestamp_allowed else '',
                file_name = '',         # Save filename         TODO: Get filename
                file_extension = ''     # Filename extension    TODO: Get extension
                )

            return save_filename
        
        pass
    
    def reset_workbook(self):
        pass

    def assert_worksheet_exists(self, assert_worksheet_name: str):
        assertion_result = False 
        if assert_worksheet_name in self._active_workbook_worksheet_list:
            assertion_result = True
        return assertion_result

    def create_worksheet(self, create_worksheet_name: str, create_index: int = 1):
        self._active_workbook.create_sheet(title=create_worksheet_name, index=create_index)

    def switch_worksheet(self, target_worksheet_name: str):
        if self.assert_worksheet_exists(target_worksheet_name):
            self._active_worksheet_name: str = target_worksheet_name
            self._active_worksheet: openpyxl.Workbook[str] = self._active_workbook[target_worksheet_name]
        else:
            raise NameError

    def read_parameter_worksheet(self):

        def read_parameter_string(parameter_string: str):
            parameter_string_split = parameter_string.split(',,')
            parameter_pair_list = [parameter_pair.split('=') for parameter_pair in parameter_string_split]
            return parameter_pair_list
        
        # Switching worksheet:
        parameter_worksheet_name = self._default_worksheet_name_param
        self.switch_worksheet(target_worksheet_name=parameter_worksheet_name)
    
        # Parameter list variable:
        parameter_list = []

        # Reading parameters:
        row_max = self._active_worksheet.max_row
        row_range = range(1, row_max + 1)
        column = 'A'
        for row in row_range:
            cell_target = f'{column}{row}'
            cell_value = self.read_cell(cell_position=cell_target)
            if cell_value is not None:
                parameter_string = cell_value
                parameter_pair_list = read_parameter_string(parameter_string)
                parameter_list.append(parameter_pair_list)
        
    def get_cell(self, cell_position: str):
        cell_value = self.read_cell(cell_position)
        cell_object = Cell(position=cell_position, value=cell_value)
        return cell_object
    
    def read_cell(self, cell_position: str):
        cell_value = self._active_worksheet[cell_position].value
        return cell_value
    
    def write_cell(self, cell_position: str, write_value: Any):
        self._active_worksheet[cell_position] = write_value

    def get_header_column(self, target_worksheet_name: str, header_value: str):

        # Switching worksheets:
        self.switch_worksheet(target_worksheet_name)

        # Getting all columns:
        column_list = self.get_column_list()

        # Reading headers:
        row: int = self._reserved_row
        for column in column_list:
            cell_header = f'{column}{row}'
            cell_header_value = self.read_cell(cell_header)
            if cell_header_value == header_value:
                return column
            elif cell_header_value is None:
                return column



@dataclass(frozen=True, order=True)
class ParamResult:
    param_target_worksheet_name: str = None
    param_check_type: str = None
    param_check_custom_name: str = None
    param_result_info: str = None
    param_result_count: int = 0


class ParamCore:

    def __init__(self):

        self.param_type_name: str = None
        self.param_type_code: str = None

        # Core settings:
        self.param_target_worksheet_name: str = None
        self.param_check_type: str = None
        self.param_check_custom_name: str = None

        # Highlight cell settings:
        self.param_highlight_cell: bool = None
        self.param_highlight_cell_hue: str = None
        self.param_highlight_cell_hue_default: str = None
        self.param_highlight_cell_pattern: PatternFill = None

        # Flag settings:
        self.param_flag_header_name: str = None
        self.param_flag_header_name_default: str = None
        self.param_flag_header_col: str = None
        self.param_flag_value: str = None
        self.param_flag_value_default: str = None
        self.param_flag_required: bool = True

        self._validated: bool = False
        self._ready: bool = False
        self._result: ParamResult = None
        self._result_info: str = None
        self._result_count: int = 0


    def __repr__(self):
        param_repr = f'{self.param_check_custom_name} ({self.param_type_code})'
        return param_repr

    @property
    def display(self):
        display_string = '\"{param_name}\" {param_type_code} {param_info}@{param_target_worksheet}'.format(
            param_name=self.param_check_custom_name,
            param_type_code=self.param_type_code,
            param_info=f'({self._result_info}) ' if self._result_info is not None else '',
            param_target_worksheet=self.param_target_worksheet_name,
            )
        return display_string

    @property
    def validated(self):
        return self._validated

    @property
    def ready(self):
        return self._ready

    @property
    def result(self):
        return self._result

    def _validate(self):

        self._validated = True

        # CORE ATTRIBUTES:
        param_core_attributes_list = (
            self.param_target_worksheet_name,
            self.param_check_type,
            self.param_check_custom_name)
        for param_core_attribute in param_core_attributes_list:
            if param_core_attribute is None:
                self._validated = False

        # HIGHLIGHT CELL ATTRIBUTES:
        if isinstance(self.param_highlight_cell, bool):
            if self.param_highlight_cell_hue is not None:
                if not isinstance(self.param_highlight_cell_hue, str):
                    self._validated = False
                else:
                    hue_index_len = len(self.param_highlight_cell_hue)
                    if hue_index_len != 8:
                        self._validated = False
            if self.param_highlight_cell_pattern is not None:
                if not isinstance(self.param_highlight_cell_pattern, PatternFill):
                    self._validated = False
        
        else:
            self._validated = False

        # FLAG ATTRIBUTES:
        if isinstance(self.param_flag_required, bool):
            if self.param_flag_required:
                param_flag_attributes_list = (
                    self.param_flag_header_name,
                    self.param_flag_header_col,
                    self.param_flag_value)
                for param_flag_attribute in param_flag_attributes_list:
                    if param_flag_attribute is not None:
                        if not isinstance(param_flag_attribute, str):
                            self._validated = False
                    else:
                        self._validated = False
                if isinstance(self.param_flag_header_col, str):
                    if len(self.param_flag_header_col) > 2:
                        self._validated = False
                    else:
                        header_col_list = Workbook.get_column_list()
                        if self.param_flag_header_col not in header_col_list:
                            self._validated = False
                else:
                    self._validated = False
        else:
            self._validated = False             

    def _feed(self):
        param_result = ParamResult(
            param_target_worksheet_name=self.param_target_worksheet_name,
            param_check_type=self.param_check_type,
            param_check_custom_name=self.param_check_custom_name,
            param_result_info=self._result_info,
            param_result_count=self._result_count
            )
        self._result = param_result

    def setup(self, **param_settings):

        # Setting up values according to settings inputs:
        self.__dict__.update(param_settings)
        self.param_highlight_cell_pattern = PatternFill(fill_type='solid',
                                                        start_color=self.param_highlight_cell_hue,
                                                        end_color=self.param_highlight_cell_hue)
        
        self.param_flag_header_name = self.param_flag_header_name_default
        self.param_flag_value = self.param_flag_value_default
        self.param_highlight_cell = False
        self.param_highlight_cell_hue = self.param_highlight_cell_hue_default
    

class ParamDuplicateRows(ParamCore):

    def __init__(self):
        
        # Core settings:
        super().__init__()
        self.param_type_name: str = 'Param Duplicate Rows'
        self.param_type_code: str = 'PDR'

        # Default settings:
        self.param_flag_header_name_default: str = self.param_type_code
        self.param_flag_value_default: str = 'Duplicate row'
        self.param_highlight_cell_hue_default: str = '00c90404'


    def __repr__(self):
        return super().__repr__()

    def setup(self, **param_settings):
        super().setup(**param_settings)
        
        # Finalizing: 
        self._feed()

    def connect(self, workbook: Workbook):
        pass


class ParamDuplicateRowsPartial(ParamCore):

    def __init__(self):

        # Core settings:
        super().__init__()
        self.param_type_name: str = 'Param Duplicate Rows Partial'
        self.param_type_code: str = 'PDR-P'

        # Column list settings:
        self.param_column_list: list or str = None
        self.param_column_list_is_range: bool = None

        # Default settings:
        self.param_flag_header_name_default: str = self.param_type_code
        self.param_flag_value_default: str = 'Duplicate row'
        self.param_highlight_cell_hue_default: str = '00f54242'
    
    def __repr__(self):
        return super().__repr__()

    def _validate(self):
        super()._validate()

        # COLUMN LIST ATTRIBUTES:
        if isinstance(self.param_column_list, list):
            if self.param_column_list_is_range:
                column_list_global = Workbook.get_column_list()
                column_start = self.param_column_list[0]
                column_end = self.param_column_list[-1]
                column_start_index = column_list_global.index(column_start)
                column_end_index = column_list_global.index(column_end)
                column_index_difference = int(column_end_index - column_start_index)
                column_count = len(self.param_column_list)
                if column_index_difference != column_count:
                    self._validated = False
        else:
            self._validated = False

    def setup(self, **param_settings):
        super().setup(**param_settings)

        # Generating column list:
        column_list_str = str(self.param_column_list).replace(' ', '')
        column_list_formatted = column_list_str.split(',')
        if self.param_column_list_is_range:
            column_start, column_end = column_list_formatted
            column_list_global = Workbook.get_column_list()
            column_start_index = column_list_global.index(column_start)
            column_end_index = column_list_global.index(column_end)
            column_current_index = column_start_index
            column_list_formatted = []
            while column_current_index <= column_end_index:
                column = column_list_global[column_current_index]
                if column not in column_list_formatted:
                    column_list_formatted.append(column)
                column_current_index += 1
        self.param_column_list = column_list_formatted
        
        # Updating result info string line:
        column_list_str_res = ''
        column_start = self.param_column_list[0]
        column_end = self.param_column_list[-1]
        if self.param_column_list_is_range:
            column_list_str_res = f'{column_start}:{column_end}'
        else:
            column_end = self.param_column_list[-1]
            for column in self.param_column_list:
                if column_list_str_res == '':
                    column_list_str_res = f'{column}'
                else:
                    column_list_str_res = f'{column_list_str_res}, {column}'
        result_info_str = f'{column_list_str_res}'
        self._result_info = result_info_str
        
        # Finalizing: 
        self._feed()

    def connect(self, workbook: Workbook):
        pass


class ParamEmptyCells(ParamCore):

    def __init__(self):

        # Core settings:
        super().__init__()
        self.param_type_name: str = 'Param Empty Cells'
        self.param_type_code: str = 'PEC'

        # Column list settings:
        self.param_column_list: list or str = None
        self.param_column_list_is_range: bool = None

        # Default settings:
        self.param_flag_header_name_default: str = self.param_type_code
        self.param_flag_value_default: str = 'Empty cell'
        self.param_highlight_cell_hue_default: str = '00e6cd45'
    
    def __repr__(self):
        return super().__repr__()

    def _validate(self):
        super()._validate()

        # COLUMN LIST ATTRIBUTES:
        if isinstance(self.param_column_list, list):
            if self.param_column_list_is_range:
                column_list_global = Workbook.get_column_list()
                column_start = self.param_column_list[0]
                column_end = self.param_column_list[-1]
                column_start_index = column_list_global.index(column_start)
                column_end_index = column_list_global.index(column_end)
                column_index_difference = int((column_end_index + 1) - column_start_index)
                column_count = len(self.param_column_list)
                if column_index_difference != column_count:
                    self._validated = False
        else:
            self._validated = False

    def setup(self, **param_settings):
        super().setup(**param_settings)

        # Generating column list:
        column_list_str = str(self.param_column_list).replace(' ', '')
        column_list_formatted = column_list_str.split(',')
        if self.param_column_list_is_range:
            column_start, column_end = column_list_formatted
            column_list_global = Workbook.get_column_list()
            column_start_index = column_list_global.index(column_start)
            column_end_index = column_list_global.index(column_end)
            column_current_index = column_start_index
            column_list_formatted = []
            while column_current_index <= column_end_index:
                column = column_list_global[column_current_index]
                if column not in column_list_formatted:
                    column_list_formatted.append(column)
                column_current_index += 1
        self.param_column_list = column_list_formatted
        
        # Updating result info string line:
        column_list_str_res = ''
        column_start = self.param_column_list[0]
        column_end = self.param_column_list[-1]
        if self.param_column_list_is_range:
            column_list_str_res = f'{column_start}:{column_end}'
        else:
            column_end = self.param_column_list[-1]
            for column in self.param_column_list:
                if column_list_str_res == '':
                    column_list_str_res = f'{column}'
                else:
                    column_list_str_res = f'{column_list_str_res}, {column}'
        result_info_str = f'{column_list_str_res}'
        self._result_info = result_info_str
        
        # Finalizing: 
        self._feed()

    def connect(self, workbook: Workbook):
        pass


class ParamCompareFlats(ParamCore):
    
    def __init__(self):
        
        # Core settings:
        super().__init__()

        # Type name and code:
        self.param_type_name: str = 'Param Compare Flats'
        self.param_type_code: str = 'PCF'

        # Column list settings:
        self.param_column_list: list or str = None
        self.param_column_list_is_range: bool = None

        # Value settings:
        self.param_compare_value: list or float = None
        self.param_compare_axis: str = None
        self.param_compare_axis_inclusive: bool = False
        self.param_compare_operator: str = None

        # Default settings:
        self.param_flag_header_name_default: str = self.param_type_code
        self.param_flag_value_default: str = 'Flat offset'
        self.param_highlight_cell_hue_default: str = '001daecf'
    
    def __repr__(self):
        return super().__repr__()

    def _validate(self):
        super()._validate()

        # COLUMN LIST ATTRIBUTES:
        if isinstance(self.param_column_list, list):
            if self.param_column_list_is_range:
                column_list_global = Workbook.get_column_list()
                column_start = self.param_column_list[0]
                column_end = self.param_column_list[-1]
                column_start_index = column_list_global.index(column_start)
                column_end_index = column_list_global.index(column_end)
                column_index_difference = int((column_end_index + 1) - column_start_index)
                column_count = len(self.param_column_list)
                if column_index_difference != column_count:
                    self._validated = False
        else:
            self._validated = False
        
        # VALUE LIST ATTRIBUTE:
        if not isinstance(self.param_compare_value, float):
            self._validated = False
        
        # AXIS ATTRIBUTES:
        if isinstance(self.param_compare_axis, str):
            valid_axis_mod_list = ('more', 'less')
            if self.param_compare_axis not in valid_axis_mod_list:
                self._validated = False
        else:
            self._validated = False
        if not isinstance(self.param_compare_axis_inclusive, bool): 
            self._validated = False
        if isinstance(self.param_compare_operator, str):
            valid_operators_list = ('>', '<', '>=', '<=')
            if self.param_compare_operator not in valid_operators_list:
                self._validated = False
        else:
            self._validated = False
            

    def setup(self, **param_settings):
        super().setup(**param_settings)

        # Generating column list:
        column_list_str = str(self.param_column_list).replace(' ', '')
        column_list_formatted = column_list_str.split(',')
        if self.param_column_list_is_range:
            column_start, column_end = column_list_formatted
            column_list_global = Workbook.get_column_list()
            column_start_index = column_list_global.index(column_start)
            column_end_index = column_list_global.index(column_end)
            column_current_index = column_start_index
            column_list_formatted = []
            while column_current_index <= column_end_index:
                column = column_list_global[column_current_index]
                if column not in column_list_formatted:
                    column_list_formatted.append(column)
                column_current_index += 1
        self.param_column_list = column_list_formatted
            
        # Updating result info string line:
        column_list_str_res = ''
        column_start = self.param_column_list[0]
        column_end = self.param_column_list[-1]
        if self.param_column_list_is_range:
            column_list_str_res = f'{column_start}:{column_end}'
        else:
            column_end = self.param_column_list[-1]
            for column in self.param_column_list:
                if column_list_str_res == '':
                    column_list_str_res = f'{column}'
                else:
                    column_list_str_res = f'{column_list_str_res}, {column}'
        result_info_precon = '{column_range} & {operator}{value}'.format(
            column_range=column_list_str_res,
            operator=self.param_compare_operator,
            value=self.param_compare_value)
        self._result_info = result_info_precon
        
        # Finalizing: 
        self._feed()

    def connect(self, workbook: Workbook):
        pass


class ParamCompareSums(ParamCore):
    
    def __init__(self):

        # Core settings:
        super().__init__()
        self.param_type_name: str = 'Param Compare Sums'
        self.param_type_code: str = 'PCS'

        # Column list settings:
        self.param_column_list: list or str = None
        self.param_column_sum: list or str = None

        # Default settings:
        self.param_flag_header_name_default: str = self.param_type_code
        self.param_flag_value_default: str = 'Sum offset'
        self.param_highlight_cell_hue_default: str = '00ff61ab'
    
    def __repr__(self):
        return super().__repr__()
    
    def _validate(self):
        super()._validate()

        # COLUMN LIST ATTRIBUTES:
        if not isinstance(self.param_column_list, list):
            self._validated = False
    
    def setup(self, **param_settings):
        super().setup(**param_settings)

        # Generating column list:
        column_list_str = str(self.param_column_list).replace(' ', '')
        column_list_formatted = column_list_str.split(',')
        self.param_column_list = column_list_formatted
        column_sum_str = str(self.param_column_sum).replace(' ', '')
        column_sum_str_formatted = column_sum_str.split(',')
        self.param_column_sum = column_sum_str_formatted

        # Updating result info string line:
        column_list_str_res = ''
        column_add_list = self.param_column_list[:-1]
        for column in column_add_list:
            if column_add_list.index(column) == 0:
                column_list_str_res = f'{column}'
            else:
                column_list_str_res = f'{column_list_str_res} + {column}'
        column_sum_check = self.param_column_list[-1]
        column_list_str_res = f'{column_list_str_res} <> {column_sum_check}'
        result_info_str = f'{column_list_str_res}'
        self._result_info = result_info_str

        # Finalizing: 
        self._feed()


class ParamCompareTime(ParamCore):
    pass
 

@dataclass(frozen=True, order=True)
class ParamListed:
    name_displayed: str = None
    parameter_class_object: ParamCore = None

class AppWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        self.app_workbook = Workbook()
        self.app_window_title: str = 'Test Interface'
        self.app_layout: Any = None

        self.parameters_list = []

    
    def setup(self):

        def create_button(button_caption: str):
            button_object = QPushButton()
            button_object.setText(button_caption)
            button_object.setFont(QFont('DengXian', 12))
            button_width = 110
            button_height = 40
            button_object.setFixedSize(button_width, button_height)
            return button_object

        layout_grid = QGridLayout()

        
        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Open", main menu:
        def button_open_event():
            filename_bundle = QFileDialog.getOpenFileName(
                self, caption='Open file',
                directory='',
                filter='Excel Document (*.xlsx)'
                )
            
            # Getting file path:
            filename_path = filename_bundle[0]
            if len(filename_path) > 0:

                # Opening workbook:
                self.app_workbook.open_workbook(document_path=filename_path)

                # Updating main menu buttons status:
                button_open.setDisabled(True)
                button_save.setDisabled(False)
                button_close.setDisabled(False)

                # Updating parameter manager buttons status:
                button_add.setDisabled(False)

                # Checking if workbook contains P_LIST worksheet, enabling read option:
                default_param_worksheet_name = self.app_workbook._default_worksheet_name_param
                if default_param_worksheet_name in self.app_workbook._active_workbook_worksheet_list:
                    button_read.setDisabled(False)
            
        button_open_caption = 'Open'
        button_open = create_button(button_caption=button_open_caption)
        button_open.clicked.connect(lambda: button_open_event())
        button_open.setDisabled(False)
        layout_grid.addWidget(button_open, 0, 0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Save", main menu:
        def button_save_event():

            # Saving workbook:
            self.app_workbook.save_workbook()
            pass

        button_save_caption = 'Save'
        button_save = create_button(button_caption=button_save_caption)
        button_save.clicked.connect(lambda: button_save_event())
        button_save.setDisabled(True)
        layout_grid.addWidget(button_save, 0, 1)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Close", main menu:
        def button_close_event():

            # Closing workbook:
            self.app_workbook.close_workbook()

            # Updating main menu buttons status:
            button_open.setDisabled(False)
            button_save.setDisabled(True)
            button_close.setDisabled(True)

            # Updating parameter manager buttons status:
            button_read.setDisabled(True)
            button_add.setDisabled(True)
            button_edit.setDisabled(True)

        button_close_caption = 'Close'
        button_close = create_button(button_caption=button_close_caption)
        button_close.clicked.connect(lambda: button_close_event())
        button_close.setDisabled(True)
        layout_grid.addWidget(button_close, 0, 2)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Start", main menu:
        def button_start_event():
            pass

        button_start_caption = 'Start'
        button_start = create_button(button_caption=button_start_caption)
        button_start.clicked.connect(lambda: button_start_event())
        button_start.setDisabled(True)
        layout_grid.addWidget(button_start, 0, 3)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Separator:
        
        label_parameters = QLabel()
        label_parameters_text = f''
        label_parameters.setText(label_parameters_text)
        label_parameters.setFont(QFont('DengXian', 12))
        label_parameters.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        layout_grid.addWidget(label_parameters, 1, 0, 1, 3)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # List "Parameters":

        def list_has_selected_items():
            if len(list_parameters.selectedItems()) > 0: 
                button_remove.setDisabled(False)
                if not button_open.isEnabled():
                    button_edit.setDisabled(False)
            else: 
                button_remove.setDisabled(True)
                button_edit.setDisabled(True)

        list_parameters = QListWidget()
        list_parameters.itemSelectionChanged.connect(lambda: list_has_selected_items())
        layout_grid.addWidget(list_parameters, 2, 1, 5, 3)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Read", parameters manager menu:

        def button_read_event():
            
            # Reading parameter setups from P_LIST in workbook:
            
                self.app_workbook.read_parameter_worksheet()

        button_read_caption = 'Read'
        button_read = create_button(button_caption=button_read_caption)
        button_read.clicked.connect(lambda: button_read_event())
        button_read.setDisabled(True)
        layout_grid.addWidget(button_read, 2, 0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Add", parameters manager menu:

        def button_add_event():

            def load_add_parameter_widgets():

                # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                # Disabling main and side menu buttons:
                button_save.setDisabled(True)           # Main menu
                button_close.setDisabled(True)
                button_start.setDisabled(True)
                button_read.setDisabled(True)           # Side menu
                button_add.setDisabled(True)
                button_remove.setDisabled(True)
                button_edit.setDisabled(True)
                button_save_param.setDisabled(True)
                list_parameters.setDisabled(False)

                # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                # Core settings:

                label_new_parameter = QLabel()
                label_new_parameter_text = f'Add new parameter check'
                label_new_parameter.setText(label_new_parameter_text)
                label_new_parameter.setFont(QFont('DengXian', 12))
                label_new_parameter.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
                layout_grid.addWidget(label_new_parameter, 8, 0, 1, 4)

                label_new_parameter_name = QLabel()
                label_new_parameter_name_text = 'Name'
                label_new_parameter_name.setText(label_new_parameter_name_text)
                label_new_parameter_name.setFont(QFont('DengXian', 12))
                label_new_parameter_name.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                layout_grid.addWidget(label_new_parameter_name, 9, 0)

                textbox_new_parameter_name = QLineEdit()
                textbox_new_parameter_name_text = ''
                textbox_new_parameter_name.setText(textbox_new_parameter_name_text)
                textbox_new_parameter_name.setFont(QFont('DengXian', 12))
                textbox_new_parameter_name.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                layout_grid.addWidget(textbox_new_parameter_name, 9, 1, 1, 3)

                label_new_parameter_target_worksheet = QLabel()
                label_new_parameter_target_worksheet_text = 'Worksheet'
                label_new_parameter_target_worksheet.setText(label_new_parameter_target_worksheet_text)
                label_new_parameter_target_worksheet.setFont(QFont('DengXian', 12))
                label_new_parameter_target_worksheet.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                layout_grid.addWidget(label_new_parameter_target_worksheet, 10, 0)

                dropdown_new_parameter_target_worksheet = QComboBox()
                dropdown_new_parameter_target_worksheet.setFont(QFont('DengXian', 12))
                layout_grid.addWidget(dropdown_new_parameter_target_worksheet, 10, 1, 1, 3)

                label_new_parameter_type = QLabel()
                label_new_parameter_type_text = 'Type'
                label_new_parameter_type.setText(label_new_parameter_type_text)
                label_new_parameter_type.setFont(QFont('DengXian', 12))
                label_new_parameter_type.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                layout_grid.addWidget(label_new_parameter_type, 11, 0)

                dropdown_new_parameter_type = QComboBox()
                dropdown_new_parameter_type.setFont(QFont('DengXian', 12))
                layout_grid.addWidget(dropdown_new_parameter_type, 11, 1, 1, 3)

                def button_new_parameter_core_check_event():

                    # Disabling until checked, connecting buttons:
                    button_new_parameter_continue.setDisabled(True)
                    textbox_new_parameter_name.textChanged.connect(lambda: button_new_parameter_continue.setDisabled(True))

                    # Checking if name is not empty and is valid:
                    new_parameter_name = textbox_new_parameter_name.text() 
                    if len(new_parameter_name) == 0:
                        button_new_parameter_continue.setDisabled(True)
                    else:
                        for parameter_object in self.parameters_list:
                            parameter_object_name = parameter_object.param_check_custom_name
                            if new_parameter_name == parameter_object_name:
                                button_new_parameter_continue.setDisabled(True)
                                break
                        else:
                            button_new_parameter_continue.setDisabled(False)

                button_new_parameter_check_caption = 'Check'
                button_new_parameter_check = create_button(button_caption=button_new_parameter_check_caption)
                button_new_parameter_check.clicked.connect(lambda: button_new_parameter_core_check_event())
                button_new_parameter_check.setDisabled(False)
                layout_grid.addWidget(button_new_parameter_check, 12, 1)

                def button_new_parameter_core_cancel_event():

                    # Re-enabling buttons:
                    button_save.setDisabled(False)           # Main menu
                    button_close.setDisabled(False)
                    button_read.setDisabled(False)           # Side menu
                    button_add.setDisabled(False)
                    list_parameters.setDisabled(False)
                    
                    # Disable visibility for parameter core settings UI elements:
                    for ui_element in new_parameter_core_widget_list:
                        ui_element.setVisible(False)
                        layout_grid.removeWidget(ui_element)

                button_new_parameter_cancel_caption = 'Cancel'
                button_new_parameter_cancel = create_button(button_caption=button_new_parameter_cancel_caption)
                button_new_parameter_cancel.clicked.connect(lambda: button_new_parameter_core_cancel_event())
                button_new_parameter_cancel.setDisabled(False)
                layout_grid.addWidget(button_new_parameter_cancel, 12, 2)

                def button_new_parameter_core_continue_event():

                    # Disable user input for parameter core settings:
                    for ui_element in new_parameter_core_widget_list:
                        ui_element_type_ignore_list = [QPushButton, QLabel]
                        ui_element_set_disabled = True
                        for ui_element_type in ui_element_type_ignore_list:
                            if isinstance(ui_element, ui_element_type):
                                ui_element_set_disabled = False
                                break
                        if ui_element_set_disabled:
                            ui_element.setDisabled(True)

                    # Empty function placeholders:
                    selected_parameter_type_string = dropdown_new_parameter_type.currentText()
                    selected_parameter_settings_string = {}
                    selected_parameter_object: ParamCore = None
                    button_shift_row = 0
                    def update_selected_parameter_settings(): pass

                    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    # PDR-P settings: 
                    if ParamDuplicateRowsPartial().param_type_code in selected_parameter_type_string:

                        # Preconstructing settings:
                        selected_parameter_object = ParamDuplicateRowsPartial()
                        selected_parameter_settings = {
                            'param_check_type': 'ParamDuplicateRowsPartial',
                            'param_check_custom_name': textbox_new_parameter_name.text(),
                            'param_target_worksheet_name': dropdown_new_parameter_target_worksheet.currentText(),
                            'param_column_list': [],
                            'param_column_list_is_range': False,
                            }

                        button_shift_row = 14

                        def update_selected_parameter_settings():
                            column_list_string = textbox_pdrp_column_list.text()
                            column_list_is_range = True
                            if dropdown_pdrp_column_list_is_range.currentText() == 'False':
                                 column_list_is_range = False
                            selected_parameter_settings['param_column_list'] = column_list_string
                            selected_parameter_settings['param_column_list_is_range'] = column_list_is_range

                        def check_column_input():
                            input_string = textbox_pdrp_column_list.text()
                            valid_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                            valid_separator = ','
                            if input_string == '':
                                pass
                            else:
                                if len(input_string) == 1:
                                    if input_string[0] not in valid_characters: input_string = ''
                                    elif input_string[0] == valid_separator: input_string = ''
                                else:
                                    if input_string[-1] == valid_separator:
                                        if input_string[-2] == valid_separator:
                                            input_string = input_string[:-1]
                                    elif input_string[-1] not in valid_characters:
                                        input_string = input_string[:-1]

                            prev_input = ''
                            for character in input_string:
                                if character == valid_separator:
                                    if len(prev_input) in (1, 2):
                                        prev_input = ''
                                    else:
                                        prev_input += character
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break
                                else:
                                    prev_input += character
                                    if len(prev_input) > 2:
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break

                            textbox_pdrp_column_list.setText(input_string)
                            button_new_settings_add.setEnabled(False)

                            if input_string.count(',') > 1:
                                dropdown_pdrp_column_list_is_range.setCurrentIndex(1)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Columns list input textbox --> "Columns": [__________]
                        label_pdrp_column_list = QLabel()
                        label_pdrp_column_list_text = 'Columns'
                        label_pdrp_column_list.setText(label_pdrp_column_list_text)
                        label_pdrp_column_list.setFont(QFont('DengXian', 12))
                        label_pdrp_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pdrp_column_list, 12, 0)
                        new_parameter_settings_widget_list.append(label_pdrp_column_list)

                        textbox_pdrp_column_list = QLineEdit()
                        textbox_pdrp_column_list_text = ''
                        textbox_pdrp_column_list.setText(textbox_pdrp_column_list_text)
                        textbox_pdrp_column_list.setFont(QFont('DengXian', 12))
                        textbox_pdrp_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        textbox_pdrp_column_list.textChanged.connect(lambda: check_column_input())
                        layout_grid.addWidget(textbox_pdrp_column_list, 12, 1, 1, 3)
                        new_parameter_settings_widget_list.append(textbox_pdrp_column_list)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Column list is range bool dropdown --> "Is range": [True______] 
                        #                                                    [False_____]
                        label_pdrp_column_list_is_range = QLabel()
                        label_pdrp_column_list_is_range_text = 'Is range'
                        label_pdrp_column_list_is_range.setText(label_pdrp_column_list_is_range_text)
                        label_pdrp_column_list_is_range.setFont(QFont('DengXian', 12))
                        label_pdrp_column_list_is_range.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pdrp_column_list_is_range, 13, 0)
                        new_parameter_settings_widget_list.append(label_pdrp_column_list_is_range)

                        dropdown_pdrp_column_list_is_range = QComboBox()
                        dropdown_pdrp_column_list_is_range.setFont((QFont('DengXian', 12)))
                        dropdown_pdrp_column_list_is_range.addItems(('True', 'False'))
                        dropdown_pdrp_column_list_is_range.currentIndexChanged.connect(lambda: button_new_settings_add.setDisabled(True))
                        layout_grid.addWidget(dropdown_pdrp_column_list_is_range, 13, 1, 1, 3)
                        new_parameter_settings_widget_list.append(dropdown_pdrp_column_list_is_range)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Button "Check" settings input --> [Check] [Clear] [Add] 
                        #                                      ^
                        def button_new_settings_check_event():

                            # Disabling "Add" button upon next edit:
                            textbox_pdrp_column_list.textChanged.connect(lambda: button_new_settings_add.setDisabled(True))

                            # Checking input and enabling "Add" button if input is valid::
                            input_is_valid = True
                            input_has_invalid_character = False
                            input_string = textbox_pdrp_column_list.text().upper()
                            
                            if len(input_string) > 0:
                                while input_string[-1] in (',', ' '):
                                    input_string = input_string[:-1]
                            
                            if len(input_string) == 0:
                                input_is_valid = False 
                            else:

                                # Checking invalid character input:
                                input_string_test = input_string.replace(',', '')
                                if len(input_string_test) == 0:
                                    input_is_valid = False
                                character_list_invalid = '1234567890!@#$%^&*()_+-=[]{}\|\\;\'\:\"./<>?~'
                                character_list_valid = ',abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                                for character in character_list_invalid:
                                    if character in input_string:
                                        input_has_invalid_character = True
                                        input_is_valid = False
                                for character in input_string:
                                    if character not in character_list_valid:
                                        input_has_invalid_character = True
                                        input_is_valid = False

                                if not input_has_invalid_character:
                                    column_list_is_range = False
                                    if dropdown_pdrp_column_list_is_range.currentText() == 'True':
                                        column_list_is_range = True
                    
                                    # Generating column list:
                                    column_list_str = str(input_string).replace(' ', '')
                                    if column_list_str[-1] == ',':
                                        column_list_str = column_list_str[0:-1]
                                    column_list_formatted = column_list_str.split(',')
                                    column_list = []
                                    if len(column_list_formatted) == 1:
                                        dropdown_pdrp_column_list_is_range.setCurrentIndex(1)
                                        column_list_is_range = False
                                    else:
                                        if column_list_is_range:
                                            try:
                                                column_start, column_end = column_list_formatted
                                                column_list_global = Workbook.get_column_list()
                                                column_start_index = column_list_global.index(column_start)
                                                column_end_index = column_list_global.index(column_end)
                                                column_current_index = column_start_index
                                                column_list_formatted = []
                                                while column_current_index <= column_end_index:
                                                    column = column_list_global[column_current_index]
                                                    if column not in column_list_formatted:
                                                        column_list_formatted.append(column)
                                                    column_current_index += 1
                                            except:
                                                column_list_formatted = None
                                        column_list = column_list_formatted

                                    # Generating range, if column list is range:
                                    if isinstance(column_list, list):
                                        if column_list_is_range:
                                            column_list_global = Workbook.get_column_list()
                                            column_start = column_list[0]
                                            column_end = column_list[-1]
                                            if column_start == column_end:
                                                dropdown_pdrp_column_list_is_range.setCurrentIndex(1)
                                                column_list_is_range = False
                                            else:
                                                column_start_index = column_list_global.index(column_start)
                                                column_end_index = column_list_global.index(column_end)
                                                column_index_difference = int((column_end_index + 1) - column_start_index)
                                                column_count = len(column_list)
                                                if column_index_difference != column_count:
                                                    input_is_valid = False
                                    else:
                                        input_is_valid = False 
                            
                            if input_is_valid:
                                textbox_pdrp_column_list.setText(input_string.upper())
                                button_new_settings_add.setDisabled(False)
                            else:
                                button_new_settings_add.setDisabled(True)

                        button_new_settings_check_caption = 'Check'
                        button_new_settings_check = create_button(button_caption=button_new_settings_check_caption)
                        button_new_settings_check.clicked.connect(lambda: button_new_settings_check_event())
                        button_new_settings_check.setDisabled(False)
                        layout_grid.addWidget(button_new_settings_check, button_shift_row, 1)
                        new_parameter_settings_widget_list.append(button_new_settings_check)

                    # PDR settings widgets:
                    elif ParamDuplicateRows().param_type_code in selected_parameter_type_string:
                        selected_parameter_object = ParamDuplicateRows()
                        selected_parameter_settings = {
                            'param_check_type': 'ParamDuplicateRows',
                            'param_check_custom_name': textbox_new_parameter_name.text(),
                            'param_target_worksheet_name': dropdown_new_parameter_target_worksheet.currentText(),
                            }
                        button_shift_row = 12

                    # PEC settings widgets:
                    elif ParamEmptyCells().param_type_code in selected_parameter_type_string:
                        
                        button_shift_row = 14

                        # Preconstructing settings:
                        selected_parameter_object = ParamEmptyCells()
                        selected_parameter_settings = {
                            'param_check_type': 'ParamEmptyCells',
                            'param_check_custom_name': textbox_new_parameter_name.text(),
                            'param_target_worksheet_name': dropdown_new_parameter_target_worksheet.currentText(),
                            'param_column_list': [],
                            'param_column_list_is_range': False,
                            }

                        def update_selected_parameter_settings():
                            column_list_string = textbox_pec_column_list.text()
                            column_list_is_range = True
                            if dropdown_pec_column_list_is_range.currentText() == 'False':
                                 column_list_is_range = False
                            selected_parameter_settings['param_column_list'] = column_list_string
                            selected_parameter_settings['param_column_list_is_range'] = column_list_is_range

                        def check_column_input():
                            input_string = textbox_pec_column_list.text()
                            valid_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                            valid_separator = ','
                            if input_string == '':
                                pass
                            else:
                                if len(input_string) == 1:
                                    if input_string[0] not in valid_characters: input_string = ''
                                    elif input_string[0] == valid_separator: input_string = ''
                                else:
                                    if input_string[-1] == valid_separator:
                                        if input_string[-2] == valid_separator:
                                            input_string = input_string[:-1]
                                    elif input_string[-1] not in valid_characters:
                                        input_string = input_string[:-1]

                            prev_input = ''
                            for character in input_string:
                                if character == valid_separator:
                                    if len(prev_input) in (1, 2):
                                        prev_input = ''
                                    else:
                                        prev_input += character
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break
                                else:
                                    prev_input += character
                                    if len(prev_input) > 2:
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break

                            textbox_pec_column_list.setText(input_string)
                            button_new_settings_add.setEnabled(False)

                            if input_string.count(',') > 1:
                                dropdown_pec_column_list_is_range.setCurrentIndex(1)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Columns list input textbox --> "Columns": [__________]
                        label_pec_column_list = QLabel()
                        label_pec_column_list_text = 'Columns'
                        label_pec_column_list.setText(label_pec_column_list_text)
                        label_pec_column_list.setFont(QFont('DengXian', 12))
                        label_pec_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pec_column_list, 12, 0)
                        new_parameter_settings_widget_list.append(label_pec_column_list)

                        textbox_pec_column_list = QLineEdit()
                        textbox_pec_column_list_text = ''
                        textbox_pec_column_list.setText(textbox_pec_column_list_text)
                        textbox_pec_column_list.setFont(QFont('DengXian', 12))
                        textbox_pec_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        textbox_pec_column_list.textChanged.connect(lambda: check_column_input())
                        layout_grid.addWidget(textbox_pec_column_list, 12, 1, 1, 3)
                        new_parameter_settings_widget_list.append(textbox_pec_column_list)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Column list is range bool dropdown --> "Is range": [True______] 
                        #                                                    [False_____]
                        label_pec_column_list_is_range = QLabel()
                        label_pec_column_list_is_range_text = 'Is range'
                        label_pec_column_list_is_range.setText(label_pec_column_list_is_range_text)
                        label_pec_column_list_is_range.setFont(QFont('DengXian', 12))
                        label_pec_column_list_is_range.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pec_column_list_is_range, 13, 0)
                        new_parameter_settings_widget_list.append(label_pec_column_list_is_range)

                        dropdown_pec_column_list_is_range = QComboBox()
                        dropdown_pec_column_list_is_range.setFont((QFont('DengXian', 12)))
                        dropdown_pec_column_list_is_range.addItems(('True', 'False'))
                        dropdown_pec_column_list_is_range.currentIndexChanged.connect(lambda: button_new_settings_add.setDisabled(True))
                        layout_grid.addWidget(dropdown_pec_column_list_is_range, 13, 1, 1, 3)
                        new_parameter_settings_widget_list.append(dropdown_pec_column_list_is_range)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Button "Check" settings input --> [Check] [Clear] [Add] 
                        #                                      ^
                        def button_new_settings_check_event():

                            # Disabling "Add" button upon next edit:
                            textbox_pec_column_list.textChanged.connect(lambda: button_new_settings_add.setDisabled(True))

                            # Checking input and enabling "Add" button if input is valid::
                            input_is_valid = True
                            input_has_invalid_character = False
                            input_string = textbox_pec_column_list.text().upper()
                            
                            if len(input_string) > 0:
                                while input_string[-1] in (',', ' '):
                                    input_string = input_string[:-1]
                            
                            if len(input_string) == 0:
                                input_is_valid = False 
                            else:

                                # Checking invalid character input:
                                input_string_test = input_string.replace(',', '')
                                if len(input_string_test) == 0:
                                    input_is_valid = False
                                character_list_invalid = '1234567890!@#$%^&*()_+-=[]{}\|\\;\'\:\"./<>?~'
                                character_list_valid = ',abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                                for character in character_list_invalid:
                                    if character in input_string:
                                        input_has_invalid_character = True
                                        input_is_valid = False
                                for character in input_string:
                                    if character not in character_list_valid:
                                        input_has_invalid_character = True
                                        input_is_valid = False

                                if not input_has_invalid_character:
                                    column_list_is_range = False
                                    if dropdown_pec_column_list_is_range.currentText() == 'True':
                                        column_list_is_range = True
                    
                                    # Generating column list:
                                    column_list_str = str(input_string).replace(' ', '')
                                    if column_list_str[-1] == ',':
                                        column_list_str = column_list_str[0:-1]
                                    column_list_formatted = column_list_str.split(',')
                                    column_list = []
                                    if len(column_list_formatted) == 1:
                                        dropdown_pec_column_list_is_range.setCurrentIndex(1)
                                        column_list_is_range = False
                                    else:
                                        if column_list_is_range:
                                            try:
                                                column_start, column_end = column_list_formatted
                                                column_list_global = Workbook.get_column_list()
                                                column_start_index = column_list_global.index(column_start)
                                                column_end_index = column_list_global.index(column_end)
                                                column_current_index = column_start_index
                                                column_list_formatted = []
                                                while column_current_index <= column_end_index:
                                                    column = column_list_global[column_current_index]
                                                    if column not in column_list_formatted:
                                                        column_list_formatted.append(column)
                                                    column_current_index += 1
                                            except:
                                                column_list_formatted = None
                                        column_list = column_list_formatted

                                    # Generating range, if column list is range:
                                    if isinstance(column_list, list):
                                        if column_list_is_range:
                                            column_list_global = Workbook.get_column_list()
                                            column_start = column_list[0]
                                            column_end = column_list[-1]
                                            if column_start == column_end:
                                                dropdown_pdrp_column_list_is_range.setCurrentIndex(1)
                                                column_list_is_range = False
                                            else:
                                                column_start_index = column_list_global.index(column_start)
                                                column_end_index = column_list_global.index(column_end)
                                                column_index_difference = int((column_end_index + 1) - column_start_index)
                                                column_count = len(column_list)
                                                if column_index_difference != column_count:
                                                    input_is_valid = False
                                    else:
                                        input_is_valid = False 
                            
                            if input_is_valid:
                                textbox_pec_column_list.setText(input_string.upper())
                                button_new_settings_add.setDisabled(False)
                            else:
                                button_new_settings_add.setDisabled(True)

                        button_new_settings_check_caption = 'Check'
                        button_new_settings_check = create_button(button_caption=button_new_settings_check_caption)
                        button_new_settings_check.clicked.connect(lambda: button_new_settings_check_event())
                        button_new_settings_check.setDisabled(False)
                        layout_grid.addWidget(button_new_settings_check, button_shift_row, 1)
                        new_parameter_settings_widget_list.append(button_new_settings_check)

                    # PCF settings widgets:
                    elif ParamCompareFlats().param_type_code in selected_parameter_type_string:
                        
                        button_shift_row = 15

                        selected_parameter_object = ParamCompareFlats()
                        selected_parameter_settings = {
                            'param_check_type': 'ParamEmptyCells',
                            'param_check_custom_name': textbox_new_parameter_name.text(),
                            'param_target_worksheet_name': dropdown_new_parameter_target_worksheet.currentText(),
                            'param_column_list': [],
                            'param_column_list_is_range': False,
                            'param_compare_value': None,
                            'param_compare_axis': None,
                            'param_compare_axis_inclusive': None,
                            'param_compare_operator': None
                            }

                        def update_selected_parameter_settings():

                            # Updating columns:
                            column_list_string = textbox_pcf_column_list.text()
                            column_list_is_range = True
                            if dropdown_pcf_column_list_is_range.currentText() == 'False':
                                 column_list_is_range = False
                            selected_parameter_settings['param_column_list'] = column_list_string
                            selected_parameter_settings['param_column_list_is_range'] = column_list_is_range

                            # Updating operators:
                            compare_value = float(textbox_pcf_value.text())
                            compare_operator = dropdown_pcf_operator.currentText()
                            compare_axis = 'More' if compare_operator in ('>', '>=') else 'Less'
                            compare_axis_inclusive = True if '=' in compare_operator else False
                            selected_parameter_settings['param_compare_value'] = compare_value
                            selected_parameter_settings['param_compare_axis'] = compare_axis
                            selected_parameter_settings['param_compare_axis_inclusive'] = compare_axis_inclusive
                            selected_parameter_settings['param_compare_operator'] = compare_operator

                        def check_column_input():
                            input_string = textbox_pcf_column_list.text()
                            valid_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                            valid_separator = ','
                            if input_string == '':
                                pass
                            else:
                                if len(input_string) == 1:
                                    if input_string[0] not in valid_characters: input_string = ''
                                    elif input_string[0] == valid_separator: input_string = ''
                                else:
                                    if input_string[-1] == valid_separator:
                                        if input_string[-2] == valid_separator:
                                            input_string = input_string[:-1]
                                    elif input_string[-1] not in valid_characters:
                                        input_string = input_string[:-1]

                            prev_input = ''
                            for character in input_string:
                                if character == valid_separator:
                                    if len(prev_input) in (1, 2):
                                        prev_input = ''
                                    else:
                                        prev_input += character
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break
                                else:
                                    prev_input += character
                                    if len(prev_input) > 2:
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break

                            textbox_pcf_column_list.setText(input_string)

                            if input_string.count(',') > 1:
                                dropdown_pcf_column_list_is_range.setCurrentIndex(1)

                        def check_value_input():
                            input_string = textbox_pcf_value.text()
                            valid_characters = '1234567890'
                            valid_separator = '.'
                            if input_string == '':
                                pass
                            else:
                                if len(input_string) == 1:
                                    if input_string[0] not in valid_characters: input_string = ''
                                    elif input_string[0] == valid_separator: input_string = ''
                                else:
                                    if input_string[-1] not in f'{valid_characters}{valid_separator}':
                                        while input_string[-1] not in f'{valid_characters}{valid_separator}':
                                            if len(input_string) == 1:
                                                if input_string[0] not in valid_characters: input_string = ''
                                                elif input_string[0] == valid_separator: input_string = ''
                                                break
                                            else: 
                                                input_string = input_string[:-1]
                                    else:
                                        if input_string[-1] == valid_separator and '.' in input_string[:-1]:
                                            while input_string[-1] == valid_separator:
                                                input_string = input_string[:-1]
                                                if len(input_string) == 2:
                                                    break
                                                else:
                                                    if input_string[-1] == valid_separator and '.' not in input_string[:-1]:
                                                        break
                            textbox_pcf_value.setText(input_string)
                            button_new_settings_add.setEnabled(False)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Columns list input textbox --> "Columns": [__________]
                        label_pcf_column_list = QLabel()
                        label_pcf_column_list_text = 'Columns'
                        label_pcf_column_list.setText(label_pcf_column_list_text)
                        label_pcf_column_list.setFont(QFont('DengXian', 12))
                        label_pcf_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pcf_column_list, 12, 0)
                        new_parameter_settings_widget_list.append(label_pcf_column_list)

                        textbox_pcf_column_list = QLineEdit()
                        textbox_pcf_column_list_text = ''
                        textbox_pcf_column_list.setText(textbox_pcf_column_list_text)
                        textbox_pcf_column_list.setFont(QFont('DengXian', 12))
                        textbox_pcf_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        textbox_pcf_column_list.textChanged.connect(lambda: check_column_input())
                        layout_grid.addWidget(textbox_pcf_column_list, 12, 1, 1, 3)
                        new_parameter_settings_widget_list.append(textbox_pcf_column_list)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Column list is range bool dropdown --> "Is range": [True______] 
                        #                                                    [False_____]
                        label_pcf_column_list_is_range = QLabel()
                        label_pcf_column_list_is_range_text = 'Is range'
                        label_pcf_column_list_is_range.setText(label_pcf_column_list_is_range_text)
                        label_pcf_column_list_is_range.setFont(QFont('DengXian', 12))
                        label_pcf_column_list_is_range.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pcf_column_list_is_range, 13, 0)
                        new_parameter_settings_widget_list.append(label_pcf_column_list_is_range)

                        dropdown_pcf_column_list_is_range = QComboBox()
                        dropdown_pcf_column_list_is_range.setFont((QFont('DengXian', 12)))
                        dropdown_pcf_column_list_is_range.addItems(('True', 'False'))
                        dropdown_pcf_column_list_is_range.currentIndexChanged.connect(lambda: button_new_settings_add.setDisabled(True))
                        layout_grid.addWidget(dropdown_pcf_column_list_is_range, 13, 1, 1, 3)
                        new_parameter_settings_widget_list.append(dropdown_pcf_column_list_is_range)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Evaluate inputs --> "Operator": [GT________]
                        #                                 [GE________]
                        #                                 [LT________]
                        #                                 [LE________]
                        label_pcf_operator = QLabel()
                        label_pcf_operator_text = 'Operator'
                        label_pcf_operator.setText(label_pcf_operator_text)
                        label_pcf_operator.setFont(QFont('DengXian', 12))
                        label_pcf_operator.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pcf_operator, 14, 0)
                        new_parameter_settings_widget_list.append(label_pcf_operator)

                        dropdown_pcf_operator = QComboBox()
                        dropdown_pcf_operator.setFont((QFont('DengXian', 12)))
                        dropdown_pcf_operator.addItems(('>', '>=', '<', '<='))
                        layout_grid.addWidget(dropdown_pcf_operator, 14, 1)
                        new_parameter_settings_widget_list.append(dropdown_pcf_operator)

                        label_pcf_value = QLabel()
                        label_pcf_value_text = 'Compare to'
                        label_pcf_value.setText(label_pcf_value_text)
                        label_pcf_value.setFont(QFont('DengXian', 12))
                        label_pcf_value.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pcf_value, 14, 2)
                        new_parameter_settings_widget_list.append(label_pcf_value)

                        textbox_pcf_value = QLineEdit()
                        textbox_pcf_value_text = ''
                        textbox_pcf_value.setText(textbox_pcf_value_text)
                        textbox_pcf_value.setFont(QFont('DengXian', 12))
                        textbox_pcf_value.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        textbox_pcf_value.textChanged.connect(lambda: check_value_input())
                        layout_grid.addWidget(textbox_pcf_value, 14, 3)
                        new_parameter_settings_widget_list.append(textbox_pcf_value)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Button "Check" settings input --> [Check] [Clear] [Add] 
                        #      
                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Button "Check" settings input --> [Check] [Clear] [Add] 
                        #                                      ^
                        def button_new_settings_check_event():

                            # Disabling "Add" button upon next edit:
                            textbox_pcf_column_list.textChanged.connect(lambda: button_new_settings_add.setDisabled(True))

                            # Checking input and enabling "Add" button if input is valid::
                            column_input_is_valid = True
                            input_has_invalid_character = False
                            input_string = textbox_pcf_column_list.text().upper()
                            
                            # Checking columns:
                            if len(input_string) > 0:
                                while input_string[-1] in (',', ' '):
                                    input_string = input_string[:-1]
                            
                            if len(input_string) == 0:
                                column_input_is_valid = False 
                            else:

                                # Checking invalid character input:
                                input_string_test = input_string.replace(',', '')
                                if len(input_string_test) == 0:
                                    column_input_is_valid = False
                                character_list_invalid = '1234567890!@#$%^&*()_+-=[]{}\|\\;\'\:\"./<>?~'
                                character_list_valid = ',abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                                for character in character_list_invalid:
                                    if character in input_string:
                                        input_has_invalid_character = True
                                        column_input_is_valid = False
                                for character in input_string:
                                    if character not in character_list_valid:
                                        input_has_invalid_character = True
                                        column_input_is_valid = False

                                if not input_has_invalid_character:
                                    column_list_is_range = False
                                    if dropdown_pcf_column_list_is_range.currentText() == 'True':
                                        column_list_is_range = True
                    
                                    # Generating column list:
                                    column_list_str = str(input_string).replace(' ', '')
                                    if column_list_str[-1] == ',':
                                        column_list_str = column_list_str[0:-1]
                                    column_list_formatted = column_list_str.split(',')
                                    column_list = []
                                    if len(column_list_formatted) == 1:
                                        dropdown_pcf_column_list_is_range.setCurrentIndex(1)
                                        column_list_is_range = False
                                    else:
                                        if column_list_is_range:
                                            try:
                                                column_start, column_end = column_list_formatted
                                                column_list_global = Workbook.get_column_list()
                                                column_start_index = column_list_global.index(column_start)
                                                column_end_index = column_list_global.index(column_end)
                                                column_current_index = column_start_index
                                                column_list_formatted = []
                                                while column_current_index <= column_end_index:
                                                    column = column_list_global[column_current_index]
                                                    if column not in column_list_formatted:
                                                        column_list_formatted.append(column)
                                                    column_current_index += 1
                                            except:
                                                column_list_formatted = None
                                        column_list = column_list_formatted

                                    # Generating range, if column list is range:
                                    if isinstance(column_list, list):
                                        if column_list_is_range:
                                            column_list_global = Workbook.get_column_list()
                                            column_start = column_list[0]
                                            column_end = column_list[-1]
                                            if column_start == column_end:
                                                dropdown_pdrp_column_list_is_range.setCurrentIndex(1)
                                                column_list_is_range = False
                                            else:
                                                column_start_index = column_list_global.index(column_start)
                                                column_end_index = column_list_global.index(column_end)
                                                column_index_difference = int((column_end_index + 1) - column_start_index)
                                                column_count = len(column_list)
                                                if column_index_difference != column_count:
                                                    column_input_is_valid = False
                                    else:
                                        column_input_is_valid = False 
                            
                            # Checking compare to value:
                            compare_to_value_is_valid = True
                            compare_to_value = textbox_pcf_value.text()
                            if len(compare_to_value) > 0:
                                try:
                                    compare_to_value = float(compare_to_value)
                                    textbox_pcf_value.setText(str(compare_to_value))
                                except:
                                    pass
                            else:
                                compare_to_value_is_valid = False
                            
                            if column_input_is_valid:
                                textbox_pcf_column_list.setText(input_string.upper())
                                if compare_to_value_is_valid:
                                    button_new_settings_add.setDisabled(False)
                            else:
                                button_new_settings_add.setDisabled(True)
                            
                        button_new_settings_check_caption = 'Check'
                        button_new_settings_check = create_button(button_caption=button_new_settings_check_caption)
                        button_new_settings_check.clicked.connect(lambda: button_new_settings_check_event())
                        button_new_settings_check.setDisabled(False)
                        layout_grid.addWidget(button_new_settings_check, button_shift_row, 1)
                        new_parameter_settings_widget_list.append(button_new_settings_check)

                    # PCS settings widgets:
                    elif ParamCompareSums().param_type_code in selected_parameter_type_string:
                        
                        button_shift_row = 14

                        selected_parameter_object = ParamCompareFlats()
                        selected_parameter_settings = {
                            'param_check_type': 'ParamEmptyCells',
                            'param_check_custom_name': textbox_new_parameter_name.text(),
                            'param_target_worksheet_name': dropdown_new_parameter_target_worksheet.currentText(),
                            'param_column_list': [],
                            'param_column_sum': [],
                            }

                        def update_selected_parameter_settings():

                            # Updating columns:
                            column_list_string = textbox_pcf_column_list.text()
                            selected_parameter_settings['param_column_list'] = column_list_string
                            column_sum_string = textbox_pcs_sum_column_list.text()
                            selected_parameter_settings['param_column_sum'] = column_sum_string

                        def check_column_input():
                            input_string = textbox_pcs_column_list.text()
                            valid_characters = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
                            valid_separator = ','
                            if input_string == '':
                                pass
                            else:
                                if len(input_string) == 1:
                                    if input_string[0] not in valid_characters: input_string = ''
                                    elif input_string[0] == valid_separator: input_string = ''
                                else:
                                    if input_string[-1] == valid_separator:
                                        if input_string[-2] == valid_separator:
                                            input_string = input_string[:-1]
                                    elif input_string[-1] not in valid_characters:
                                        input_string = input_string[:-1]

                            prev_input = ''
                            for character in input_string:
                                if character == valid_separator:
                                    if len(prev_input) in (1, 2):
                                        prev_input = ''
                                    else:
                                        prev_input += character
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break
                                else:
                                    prev_input += character
                                    if len(prev_input) > 2:
                                        remove_index = len(prev_input) * -1
                                        input_string = input_string[:remove_index]
                                        prev_input = ''
                                        break

                            textbox_pcs_column_list.setText(input_string)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Columns list input textbox --> "Columns check": [__________]
                        label_pcs_column_list = QLabel()
                        label_pcs_column_list_text = 'Columns check'
                        label_pcs_column_list.setText(label_pcs_column_list_text)
                        label_pcs_column_list.setFont(QFont('DengXian', 12))
                        label_pcs_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pcs_column_list, 12, 0)
                        new_parameter_settings_widget_list.append(label_pcs_column_list)

                        textbox_pcs_column_list = QLineEdit()
                        textbox_pcs_column_list_text = ''
                        textbox_pcs_column_list.setText(textbox_pcs_column_list_text)
                        textbox_pcs_column_list.setFont(QFont('DengXian', 12))
                        textbox_pcs_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        textbox_pcs_column_list.textChanged.connect(lambda: check_column_input())
                        layout_grid.addWidget(textbox_pcs_column_list, 12, 1, 1, 3)
                        new_parameter_settings_widget_list.append(textbox_pcs_column_list)

                        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                        # Columns list input textbox --> "Column sum": [__________]
                        label_pcs_sum_column_list = QLabel()
                        label_pcs_sum_column_list_text = 'Column sum'
                        label_pcs_sum_column_list.setText(label_pcs_sum_column_list_text)
                        label_pcs_sum_column_list.setFont(QFont('DengXian', 12))
                        label_pcs_sum_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        layout_grid.addWidget(label_pcs_sum_column_list, 13, 0)
                        new_parameter_settings_widget_list.append(label_pcs_sum_column_list)

                        textbox_pcs_sum_column_list = QLineEdit()
                        textbox_pcs_sum_column_list_text = ''
                        textbox_pcs_sum_column_list.setText(textbox_pcs_sum_column_list_text)
                        textbox_pcs_sum_column_list.setFont(QFont('DengXian', 12))
                        textbox_pcs_sum_column_list.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
                        textbox_pcs_sum_column_list.textChanged.connect(lambda: check_column_input())
                        layout_grid.addWidget(textbox_pcs_sum_column_list, 13, 1, 1, 3)
                        new_parameter_settings_widget_list.append(textbox_pcs_sum_column_list)


                    # PCT settings widgets:
                    # elif ParamCompareTime().param_type_code in selected_parameter_type_string:
                    #     TODO:
                    #     pass

                    # Removing old buttons:
                    for ui_element in new_parameter_core_widget_list:
                        if isinstance(ui_element, QPushButton):
                            ui_element.setVisible(False)
                            layout_grid.removeWidget(ui_element)

                    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    # Button "Cancel" settings input --> [Check] [Cancel] [Add] 
                    #                                                ^
                    def button_new_settings_cancel_event():
                        
                        # Removing widgets:
                        for ui_element in new_parameter_core_widget_list:
                            ui_element.setVisible(False)
                            layout_grid.removeWidget(ui_element)
                        for ui_element in new_parameter_settings_widget_list:
                            ui_element.setVisible(False)
                            layout_grid.removeWidget(ui_element)

                        # Re-enabling buttons:
                        button_save.setDisabled(False)           # Main menu
                        button_close.setDisabled(False)
                        button_read.setDisabled(False)           # Side menu
                        button_add.setDisabled(False)
                        list_parameters.setDisabled(False)
                    
                    button_new_settings_cancel_caption = 'Cancel'
                    button_new_settings_cancel = create_button(button_caption=button_new_settings_cancel_caption)
                    button_new_settings_cancel.clicked.connect(lambda: button_new_settings_cancel_event())
                    button_new_settings_cancel.setDisabled(False)
                    layout_grid.addWidget(button_new_settings_cancel, button_shift_row, 2)
                    new_parameter_settings_widget_list.append(button_new_settings_cancel)

                    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
                    # Button "Cancel" settings input --> [Check] [Cancel] [Add] 
                    #                                                       ^
                    def button_new_settings_add_event():

                        # Updating parameter settings dictionary:
                        update_selected_parameter_settings()
                        selected_parameter_object.setup(**selected_parameter_settings)

                        # Getting header column:
                        target_worksheet: str = dropdown_new_parameter_target_worksheet.currentText()
                        header_value: str = selected_parameter_object.param_flag_header_name
                        header_column: str = self.app_workbook.get_header_column(target_worksheet, header_value)
                        selected_parameter_object.param_flag_header_col = header_column

                        # Validating:
                        selected_parameter_object._validate()

                        list_parameters.addItem(selected_parameter_object.display)
                        self.parameters_list.append(selected_parameter_object)
                        pprint(selected_parameter_object.__dict__)

                        # Removing widgets:
                        for ui_element in new_parameter_core_widget_list:
                            ui_element.setVisible(False)
                            ui_element.setEnabled(False)
                            layout_grid.removeWidget(ui_element)
                        for ui_element in new_parameter_settings_widget_list:
                            ui_element.setVisible(False)
                            ui_element.setEnabled(False)
                            layout_grid.removeWidget(ui_element)

                        # Re-enabling buttons:
                        button_save.setDisabled(False)           # Main menu
                        button_close.setDisabled(False)
                        button_read.setDisabled(False)           # Side menu
                        button_add.setDisabled(False)
                        list_parameters.setDisabled(False)

                    button_new_settings_add_caption = 'Add'
                    button_new_settings_add = create_button(button_caption=button_new_settings_add_caption)
                    button_new_settings_add.clicked.connect(lambda: button_new_settings_add_event())
                    button_new_settings_add.setEnabled(False)
                    layout_grid.addWidget(button_new_settings_add, button_shift_row, 3)
                    new_parameter_settings_widget_list.append(button_new_settings_add)

                    # Bypassing setEnabled(False), if type is PDR:
                    if ParamDuplicateRows().param_type_code in selected_parameter_type_string:
                        if ParamDuplicateRowsPartial().param_type_code not in selected_parameter_type_string:
                            button_new_settings_add.setEnabled(True)

                button_new_parameter_continue_caption = 'Continue'
                button_new_parameter_continue = create_button(button_caption=button_new_parameter_continue_caption)
                button_new_parameter_continue.clicked.connect(lambda: button_new_parameter_core_continue_event())
                button_new_parameter_continue.setDisabled(True)
                layout_grid.addWidget(button_new_parameter_continue, 12, 3)

                new_parameter_core_widget_list = [
                    label_new_parameter, 
                    label_new_parameter_name, textbox_new_parameter_name,
                    label_new_parameter_target_worksheet, dropdown_new_parameter_target_worksheet,
                    label_new_parameter_type, dropdown_new_parameter_type,
                    button_new_parameter_check, button_new_parameter_cancel, button_new_parameter_continue]
                for ui_element in new_parameter_core_widget_list:
                    ui_element.setVisible(False)
                
                # Adding active worksheet names list:
                dropdown_new_parameter_target_worksheet.addItems(self.app_workbook._active_workbook_worksheet_list)

                # Addint valid parameter type names and codes:
                param_class_object_list = [
                    ParamDuplicateRows(), 
                    ParamDuplicateRowsPartial(), 
                    ParamCompareFlats(),
                    ParamCompareSums(),
                    # ParamCompareTime(),
                    ParamEmptyCells(),
                    ]
                param_type_names_list = []
                for param_object in param_class_object_list:
                    param_type_option = f'{param_object.param_type_name} ({param_object.param_type_code})'
                    param_type_names_list.append(param_type_option)
                dropdown_new_parameter_type.addItems(param_type_names_list)
                
                # Setting UI elements as visible:
                for ui_element in new_parameter_core_widget_list:
                    ui_element.setVisible(True)

            load_add_parameter_widgets()

        button_add_caption = 'Add'
        button_add = create_button(button_caption=button_add_caption)
        button_add.clicked.connect(lambda: button_add_event())
        button_add.setDisabled(True)
        layout_grid.addWidget(button_add, 3, 0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Remove", parameters manager menu:

        def button_remove_event():
            for selected_item in list_parameters.selectedItems():

                # Removing from active parameters list:
                selected_parameter_string = list_parameters.currentItem().text()
                selected_parameter_name_pattern: str = '\"(\w+)\"'
                selected_parameter_name = re.findall(pattern=selected_parameter_name_pattern, 
                                                     string=selected_parameter_string)[0]
                selected_parameter: ParamCore = None
                for parameter_object in self.parameters_list:
                    parameter_object_name = parameter_object.param_check_custom_name
                    if parameter_object_name == selected_parameter_name:
                        selected_parameter = parameter_object
                        break
                
                if selected_parameter in self.parameters_list:
                    self.parameters_list.remove(selected_parameter)
                
                # Removing from display:
                list_parameters.takeItem(list_parameters.row(selected_item))

            if len(list_parameters.selectedItems()) == 0:
                button_remove.setDisabled(True)

        button_remove_caption = 'Remove'
        button_remove = create_button(button_caption=button_remove_caption)
        button_remove.clicked.connect(lambda: button_remove_event())
        button_remove.setDisabled(True)
        layout_grid.addWidget(button_remove, 4, 0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Edit", parametesdrs manager menu:

        def button_edit_event():
            selected_parameter_string = list_parameters.currentItem().text()
            selected_parameter_name_pattern: str = '\"(\w+)\"'
            selected_parameter_name = re.findall(pattern=selected_parameter_name_pattern, 
                                                 string=selected_parameter_string)[0]
            selected_parameter: ParamCore = None
            for parameter_object in self.parameters_list:
                parameter_object_name = parameter_object.param_check_custom_name
                if parameter_object_name == selected_parameter_name:
                    selected_parameter = parameter_object
                    break
            
            label_edit_parameter = QLabel()
            label_edit_parameter_text = f'Edit target worksheet'
            label_edit_parameter.setText(label_edit_parameter_text)
            label_edit_parameter.setFont(QFont('DengXian', 12))
            label_edit_parameter.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            layout_grid.addWidget(label_edit_parameter, 8, 0, 1, 4)

            label_edit_parameter_target_worksheet = QLabel()
            label_edit_parameter_target_worksheet_text = 'Worksheet'
            label_edit_parameter_target_worksheet.setText(label_edit_parameter_target_worksheet_text)
            label_edit_parameter_target_worksheet.setFont(QFont('DengXian', 12))
            label_edit_parameter_target_worksheet.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
            layout_grid.addWidget(label_edit_parameter_target_worksheet, 9, 0)

            dropdown_edit_parameter_target_worksheet = QComboBox()
            dropdown_edit_parameter_target_worksheet.setFont(QFont('DengXian', 12))
            dropdown_edit_parameter_target_worksheet.addItems(self.app_workbook._active_workbook_worksheet_list)
            layout_grid.addWidget(dropdown_edit_parameter_target_worksheet, 9, 1, 1, 3)

            def button_save_edit_event():

                # Removing edit tab widgets:
                for widget in edit_widget_list:
                    widget.setVisible(False)
                    layout_grid.removeWidget(widget)

                # Removing changed element from the list:
                selected_item = list_parameters.currentItem()
                list_parameters.takeItem(list_parameters.row(selected_item))

                # Adding new element to the list:
                new_target_worksheet_name = dropdown_edit_parameter_target_worksheet.currentText()
                selected_parameter.param_target_worksheet_name = new_target_worksheet_name
                selected_parameter_display = selected_parameter.display
                list_parameters.addItem(selected_parameter_display)

            button_save_edit_caption = 'Save'
            button_save_edit = create_button(button_caption=button_save_edit_caption)
            button_save_edit.clicked.connect(lambda: button_save_edit_event())
            button_save_edit.setDisabled(False)
            layout_grid.addWidget(button_save_edit, 10, 3)

            edit_widget_list = [
                label_edit_parameter,
                label_edit_parameter_target_worksheet, 
                dropdown_edit_parameter_target_worksheet,
                button_save_edit]
            
        button_edit_caption = 'Edit'
        button_edit = create_button(button_caption=button_edit_caption)
        button_edit.clicked.connect(lambda: button_edit_event())
        button_edit.setDisabled(True)
        layout_grid.addWidget(button_edit, 5, 0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Button "Save", parameters manager menu:

        def button_save_param_event():
            pass

        button_save_param_caption = 'Save'
        button_save_param = create_button(button_caption=button_save_param_caption)
        button_save_param.clicked.connect(lambda: button_edit_event())
        button_save_param.setDisabled(True)
        layout_grid.addWidget(button_save_param, 6, 0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Separator:
        
        label_parameters = QLabel()
        label_parameters_text = f''
        label_parameters.setText(label_parameters_text)
        label_parameters.setFont(QFont('DengXian', 12))
        label_parameters.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        layout_grid.addWidget(label_parameters, 7, 0, 1, 4)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # End separator:
        
        final_row = 17
        for row in range(7, final_row):
            label_parameters = QLabel()
            label_parameters_text = f''
            label_parameters.setText(label_parameters_text)
            label_parameters.setFont(QFont('DengXian', 12))
            label_parameters.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            label_parameters.setFixedHeight(20)
            layout_grid.addWidget(label_parameters, row, 0, 1, 4)
        label_parameters = QLabel()
        label_parameters_text = f'。'
        label_parameters.setText(label_parameters_text)
        label_parameters.setFont(QFont('DengXian', 12))
        label_parameters.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        layout_grid.addWidget(label_parameters, final_row, 0, 1, 4)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
        # Add new parameter settings widget list:
        new_parameter_settings_widget_list = []

        self.app_layout = layout_grid

        widget_container = QWidget()
        widget_container.setLayout(self.app_layout)
        self.setCentralWidget(widget_container)


    def run(self):
        pass



def main():
    app = QApplication(sys.argv)

    window = AppWindow()
    window.setup()
    window.show()

    app.exec()


if __name__ == '__main__':
    main()
